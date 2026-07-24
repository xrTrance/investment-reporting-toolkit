"""
SQL Management Summary Report Generator
------------------------------------------
Runs the warehouse risk-aggregation queries (query_warehouse.sql) and
formats the results into an executive summary workbook: total capital
at risk, exception concentration by type, and by asset class.
"""

import os
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_sql_summary(db_path="data/portfolio_warehouse.db",
                          output_path="output/sql_management_summary.xlsx"):
    if not os.path.exists(db_path):
        print(f"Error: {db_path} not found. Run initialize_warehouse.py first.")
        return

    conn = sqlite3.connect(db_path)

    q_summary = """
        SELECT COUNT(*) AS total_exceptions, SUM(ABS(mv_variance)) AS total_capital_at_risk
        FROM reconciliation_matrix WHERE exception_type != 'MATCHED';
    """
    q_by_type = """
        SELECT exception_type, COUNT(*) AS incident_count, SUM(ABS(mv_variance)) AS concentrated_risk
        FROM reconciliation_matrix WHERE exception_type != 'MATCHED'
        GROUP BY exception_type ORDER BY concentrated_risk DESC;
    """
    q_by_asset = """
        SELECT asset_class, COUNT(*) AS incident_count, SUM(ABS(mv_variance)) AS concentrated_risk
        FROM reconciliation_matrix WHERE exception_type != 'MATCHED'
        GROUP BY asset_class ORDER BY concentrated_risk DESC;
    """

    df_summary = pd.read_sql_query(q_summary, conn)
    df_by_type = pd.read_sql_query(q_by_type, conn)
    df_by_asset = pd.read_sql_query(q_by_asset, conn)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Warehouse Metrics"

    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    font_section = Font(name="Calibri", size=12, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_kpi_num = Font(name="Calibri", size=14, bold=True, color="C00000")
    font_kpi_lbl = Font(name="Calibri", size=10, bold=True, color="595959")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_kpi = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "DATA WAREHOUSE RISK ANALYTICS SUMMARY"
    ws["A1"].font = font_title
    ws["A2"] = "Source: SQLite Relational Warehouse (reconciliation_matrix)"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="595959")

    ws["A4"] = "TOTAL EXCEPTIONS"
    ws["A4"].font = font_kpi_lbl
    ws["A4"].fill = fill_kpi
    ws["A5"] = int(df_summary["total_exceptions"].iloc[0])
    ws["A5"].font = font_kpi_num
    ws["A5"].fill = fill_kpi

    ws["B4"] = "TOTAL CAPITAL AT RISK ($)"
    ws["B4"].font = font_kpi_lbl
    ws["B4"].fill = fill_kpi
    ws["B5"] = float(df_summary["total_capital_at_risk"].iloc[0])
    ws["B5"].font = font_kpi_num
    ws["B5"].fill = fill_kpi
    ws["B5"].number_format = "$#,##0.00"

    row = 8
    ws.cell(row=row, column=1, value="EXCEPTION CONCENTRATION BY TYPE").font = font_section
    row += 1
    for c_idx, h in enumerate(["Exception Type", "Incidents", "Risk Value ($)"], 1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
    row += 1
    for _, r in df_by_type.iterrows():
        ws.cell(row=row, column=1, value=r["exception_type"]).border = border
        ws.cell(row=row, column=2, value=int(r["incident_count"])).border = border
        c = ws.cell(row=row, column=3, value=float(r["concentrated_risk"]))
        c.number_format = "$#,##0.00"
        c.border = border
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="EXCEPTION CONCENTRATION BY ASSET CLASS").font = font_section
    row += 1
    for c_idx, h in enumerate(["Asset Class", "Incidents", "Risk Value ($)"], 1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
    row += 1
    for _, r in df_by_asset.iterrows():
        ws.cell(row=row, column=1, value=r["asset_class"]).border = border
        ws.cell(row=row, column=2, value=int(r["incident_count"])).border = border
        c = ws.cell(row=row, column=3, value=float(r["concentrated_risk"]))
        c.number_format = "$#,##0.00"
        c.border = border
        row += 1

    for col, width in {"A": 30, "B": 18, "C": 20}.items():
        ws.column_dimensions[col].width = width

    os.makedirs("output", exist_ok=True)
    wb.save(output_path)
    print(f"SQL management summary saved: {output_path}")


if __name__ == "__main__":
    generate_sql_summary()
