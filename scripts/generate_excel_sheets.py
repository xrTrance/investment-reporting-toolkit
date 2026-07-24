"""
Executive Excel Exception Report Generator
---------------------------------------------
Formats the day's reconciliation exceptions into an audit-ready Excel
workbook: a KPI summary block, followed by a full itemised exception
list with conditional colour-coding by exception type. Fully generic —
scales to however many positions/exceptions the reconciliation run
actually produced.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEPTION_COLOURS = {
    "CASH_BALANCE_BREAK": "F2DCDB",
    "QUANTITY_BREAK": "F2DCDB",
    "PRICING_BREAK": "FADBD8",
    "UNRECORDED_AT_CUSTODIAN": "F2DCDB",
    "UNBOOKED_INTERNAL_TRADE": "F2DCDB",
    "UNLISTED_VALUATION_LAG": "FFF2CC",
    "COUPON_TIMING_LAG": "FFF2CC",
    "CA_DRP_TIMING_LAG": "FFF2CC",
    "FX_VALUATION_BREAK": "D6EAF8",
}


def generate_report(exceptions_csv="output/exceptions_only.csv",
                     output_path="output/exception_report.xlsx"):
    if not os.path.exists(exceptions_csv):
        print(f"Error: {exceptions_csv} not found. Run reconciliation_engine.py first.")
        return

    df = pd.read_csv(exceptions_csv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Exceptions"
    ws.sheet_view.showGridLines = True

    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    font_sub = Font(name="Calibri", size=11, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    font_kpi_num = Font(name="Calibri", size=14, bold=True, color="C00000")
    font_kpi_lbl = Font(name="Calibri", size=10, bold=True, color="595959")

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_kpi = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin = Side(style="thin", color="D9D9D9")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "INVESTMENT OPERATIONS DAILY EXCEPTION REPORT"
    ws["A1"].font = font_title
    ws["A2"] = f"As-Of Date: {datetime.today().strftime('%Y-%m-%d')} | Source: Reconciliation Engine v2"
    ws["A2"].font = font_sub

    total_exceptions = len(df)
    total_risk = df["mv_variance"].abs().sum()

    ws["A4"] = "TOTAL EXCEPTIONS IDENTIFIED"
    ws["A4"].font = font_kpi_lbl
    ws["A4"].fill = fill_kpi
    ws["A5"] = total_exceptions
    ws["A5"].font = font_kpi_num
    ws["A5"].fill = fill_kpi
    ws["A5"].alignment = Alignment(horizontal="center")

    ws["B4"] = "TOTAL CAPITAL AT RISK ($)"
    ws["B4"].font = font_kpi_lbl
    ws["B4"].fill = fill_kpi
    ws["B5"] = total_risk
    ws["B5"].font = font_kpi_num
    ws["B5"].fill = fill_kpi
    ws["B5"].number_format = "$#,##0.00"
    ws["B5"].alignment = Alignment(horizontal="center")

    for col in (1, 2):
        ws.cell(row=4, column=col).border = border_cell
        ws.cell(row=5, column=col).border = border_cell

    headers = ["Security ID", "Ticker", "Security Name", "Asset Class", "Currency",
               "Qty (Internal)", "Qty (Custodian)", "Qty Variance",
               "MV (Internal)", "MV (Custodian)", "MV Variance", "Exception Type"]

    header_row = 7
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell

    row_map = {
        "Security ID": "security_id", "Ticker": "ticker", "Security Name": "security_name",
        "Asset Class": "asset_class", "Currency": "currency",
        "Qty (Internal)": "quantity_internal", "Qty (Custodian)": "quantity_custodian",
        "Qty Variance": "qty_variance", "MV (Internal)": "market_value_internal",
        "MV (Custodian)": "market_value_custodian", "MV Variance": "mv_variance",
        "Exception Type": "exception_type",
    }

    r = header_row + 1
    for _, row in df.iterrows():
        colour = EXCEPTION_COLOURS.get(row["exception_type"], "FFFFFF")
        fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        for c_idx, h in enumerate(headers, 1):
            val = row[row_map[h]]
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = font_data
            cell.fill = fill
            cell.border = border_cell
            if h in ("Qty (Internal)", "Qty (Custodian)", "Qty Variance"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif h in ("MV (Internal)", "MV (Custodian)", "MV Variance"):
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        r += 1

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=rr, column=col_idx).value or "")) for rr in range(header_row, r)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    os.makedirs("output", exist_ok=True)
    wb.save(output_path)
    print(f"Excel exception report saved: {output_path} ({total_exceptions} rows)")


if __name__ == "__main__":
    generate_report()
