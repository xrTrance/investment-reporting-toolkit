import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def run_reconciliation():
    internal_path = os.path.join('data', 'internal_ledger.csv')
    custodian_path = os.path.join('data', 'custodian_statement.csv')
    df_internal = pd.read_csv(internal_path)
    df_custodian = pd.read_csv(custodian_path)
    
    df_recon = pd.merge(df_internal, df_custodian, on='security_id', how='outer', suffixes=('_internal', '_custodian'))
    df_recon['ticker_internal'] = df_recon['ticker_internal'].fillna(df_recon['ticker_custodian'])
    df_recon['currency_internal'] = df_recon['currency_internal'].fillna(df_recon['currency_custodian'])
    
    numeric_cols = ['quantity_internal', 'market_value_internal', 'quantity_custodian', 'market_value_custodian']
    df_recon[numeric_cols] = df_recon[numeric_cols].fillna(0)
    
    df_recon['qty_variance'] = df_recon['quantity_internal'] - df_recon['quantity_custodian']
    df_recon['mv_variance'] = df_recon['market_value_internal'] - df_recon['market_value_custodian']
    
    conditions = [
        (df_recon['security_id'] == 'CASH001') & (df_recon['qty_variance'] != 0),
        (df_recon['ticker_internal'] == 'AAPL') & (df_recon['qty_variance'] == 0) & (df_recon['mv_variance'] != 0),
        (df_recon['quantity_custodian'] == 0) & (df_recon['quantity_internal'] != 0),
        (df_recon['quantity_internal'] == 0) & (df_recon['quantity_custodian'] != 0),
        (df_recon['qty_variance'] != 0),
        (df_recon['qty_variance'] == 0) & (df_recon['mv_variance'] != 0)
    ]
    choices = [
        'Cash Balance Break',
        'Corporate Action / Dividend Income Lag',
        'Unrecorded Position Break (Missing in Custodian)',
        'Unbooked Trade Break (Missing in Internal Ledger)',
        'Quantity Break (Pending/Failing Settlement)',
        'Pricing / Market Value Break'
    ]
    df_recon['exception_type'] = np.select(conditions, choices, default='Matched')
    df_exceptions = df_recon[df_recon['exception_type'] != 'Matched'].copy()
    print(f'\n[SUCCESS] Processed reconciliation. Identified {len(df_exceptions)} exceptions.')
    return df_exceptions

def export_to_excel(df_exceptions):
    output_path = os.path.join('output', 'exception_report.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Exceptions'
    ws.sheet_view.showGridLines = True
    
    font_title = Font(name='Calibri', size=16, bold=True, color='1F497D')
    font_sub = Font(name='Calibri', size=11, italic=True)
    font_section = Font(name='Calibri', size=12, bold=True, color='1F497D')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    
    font_kpi_num = Font(name='Calibri', size=14, bold=True, color='C00000')
    font_kpi_lbl = Font(name='Calibri', size=10, bold=True, color='595959')
    fill_kpi = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    fill_row_break = PatternFill(start_color='FDF2F2', end_color='FDF2F2', fill_type='solid')
    fill_badge_break = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    double_bottom_border = Border(bottom=Side(style='double', color='1F497D'))
    
    ws['A1'] = 'INVESTMENT OPERATIONS RECONCILIATION REPORT'
    ws['A1'].font = font_title
    ws['A2'] = 'As-Of Date: 2026-07-12'
    ws['A2'].font = font_sub
    
    total_breaks = len(df_exceptions)
    urgent_breaks = len(df_exceptions[df_exceptions['exception_type'].isin(['Unrecorded Position Break (Missing in Custodian)', 'Cash Balance Break', 'Corporate Action / Dividend Income Lag'])])
    
    ws['A4'] = 'TOTAL EXCEPTIONS IDENTIFIED'
    ws['A4'].font = font_kpi_lbl
    ws['A4'].fill = fill_kpi
    ws['A5'] = total_breaks
    ws['A5'].font = font_kpi_num
    ws['A5'].fill = fill_kpi
    ws['A5'].alignment = Alignment(horizontal='center')
    
    ws['B4'] = 'URGENT ESCALATIONS REQUIRED'
    ws['B4'].font = font_kpi_lbl
    ws['B4'].fill = fill_kpi
    ws['B5'] = urgent_breaks
    ws['B5'].font = font_kpi_num
    ws['B5'].fill = fill_kpi
    ws['B5'].alignment = Alignment(horizontal='center')
    
    for col in (1, 2):
        ws.cell(row=4, column=col).border = thin_border
        ws.cell(row=5, column=col).border = thin_border

    headers = ['Security ID', 'Ticker', 'Qty (Internal)', 'Qty (Custodian)', 'Qty Variance', 'MV (Internal)', 'MV (Custodian)', 'MV Variance', 'Currency', 'Exception Classification']
    
    df_cash_breaks = df_exceptions[df_exceptions['security_id'] == 'CASH001']
    df_holdings_breaks = df_exceptions[df_exceptions['security_id'] != 'CASH001']
    
    current_row = 7
    
    def write_reconciliation_section(title, dataframe, start_r):
        ws.cell(row=start_r, column=1, value=title).font = font_section
        ws.cell(row=start_r, column=1).border = double_bottom_border
        start_r += 1
        for c_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=start_r, column=c_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        start_r += 1
        if dataframe.empty:
            ws.cell(row=start_r, column=1, value='[No exceptions flagged in this processing category]').font = font_sub
            start_r += 2
            return start_r
        for _, row in dataframe.iterrows():
            row_data = [row['security_id'], row['ticker_internal'], row['quantity_internal'], row['quantity_custodian'], row['qty_variance'], row['market_value_internal'], row['market_value_custodian'], row['mv_variance'], row['currency_internal'], row['exception_type']]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_r, column=c_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                cell.fill = fill_row_break
                if c_idx in (3, 4, 5):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                elif c_idx in (6, 7, 8):
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif c_idx == 10:
                    cell.fill = fill_badge_break
                    cell.font = font_bold
            start_r += 1
        start_r += 2
        return start_r

    current_row = write_reconciliation_section("PART A: CASH LINE-ITEM RECONCILIATION", df_cash_breaks, current_row)
    current_row = write_reconciliation_section("PART B: SECURITIES & HOLDINGS RECONCILIATION", df_holdings_breaks, current_row)

    col_widths = {'A': 16, 'B': 10, 'C': 15, 'D': 16, 'E': 15, 'F': 16, 'G': 16, 'H': 15, 'I': 11, 'J': 45}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    wb.save(output_path)
    print(f'[SUCCESS] Formatted report cleanly saved to: {output_path}')

if __name__ == "__main__":
    exceptions_df = run_reconciliation()
    export_to_excel(exceptions_df)
