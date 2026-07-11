import os
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_sql_management_report():
    db_path = os.path.join('data', 'portfolio_warehouse.db')
    output_path = os.path.join('output', 'sql_management_summary.xlsx')
    
    # Connect to your local relational warehouse
    conn = sqlite3.connect(db_path)
    
    # 1. Advanced SQL Query: Total Capital At Risk
    q_summary = """
    SELECT 
        SUM(ABS(i.market_value - COALESCE(c.market_value, 0))) AS [Total_Capital_At_Risk],
        COUNT(*) AS [Total_Breaches]
    FROM internal_ledger i
    LEFT JOIN custodian_statement c ON i.security_id = c.security_id
    WHERE i.quantity != COALESCE(c.quantity, 0) OR i.market_value != COALESCE(c.market_value, 0);
    """
    
    # 2. Advanced SQL Query: Asset Risk Concentration Matrix
    q_concentration = """
    SELECT 
        i.ticker AS [Ticker],
        COUNT(*) AS [Total_Incidents],
        SUM(ABS(i.market_value - COALESCE(c.market_value, 0))) AS [Concentrated_Risk]
    FROM internal_ledger i
    LEFT JOIN custodian_statement c ON i.security_id = c.security_id
    WHERE i.quantity != COALESCE(c.quantity, 0) OR i.market_value != COALESCE(c.market_value, 0)
    GROUP BY i.ticker
    ORDER BY [Concentrated_Risk] DESC;
    """
    
    # Read the relational data straight into dataframes
    df_summary = pd.read_sql_query(q_summary, conn)
    df_concentration = pd.read_sql_query(q_concentration, conn)
    conn.close()
    
    # Initialize the openpyxl workbook layout
    wb = Workbook()
    ws = wb.active
    ws.title = 'SQL Warehouse Metrics'
    ws.sheet_view.showGridLines = True
    
    # Corporate layout typography components
    font_title = Font(name='Calibri', size=16, bold=True, color='1F497D')
    font_section = Font(name='Calibri', size=12, bold=True, color='1F497D')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    fill_kpi = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    fill_row_break = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Soft warning amber
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    # Main Report Title Headers
    ws['A1'] = 'DATA WAREHOUSE RISK ANALYTICS REPORT'
    ws['A1'].font = font_title
    ws['A2'] = 'Source Engine: SQLite Relational Matrix | Mode: Automated'
    ws['A2'].font = Font(name='Calibri', size=11, italic=True)
    
    # Write KPI Summary Cards Block (Rows 4-5)
    ws['A4'] = 'AGGREGATED CAPITAL AT RISK'
    ws['A4'].font = Font(name='Calibri', size=10, bold=True, color='595959')
    ws['A4'].fill = fill_kpi
    ws['A4'].border = thin_border
    
    # Extract absolute value safely from the dataframe
    total_risk = float(df_summary['Total_Capital_At_Risk'].iloc[0] or 0)
    ws['A5'] = total_risk
    ws['A5'].font = Font(name='Calibri', size=14, bold=True, color='C00000')
    ws['A5'].fill = fill_kpi
    ws['A5'].number_format = '0,##0.00'
    ws['A5'].border = thin_border
    
    ws['B4'] = 'TOTAL AUDIT BREACHES'
    ws['B4'].font = Font(name='Calibri', size=10, bold=True, color='595959')
    ws['B4'].fill = fill_kpi
    ws['B4'].border = thin_border
    
    total_breaches = int(df_summary['Total_Breaches'].iloc[0] or 0)
    ws['B5'] = total_breaches
    ws['B5'].font = Font(name='Calibri', size=14, bold=True, color='1F497D')
    ws['B5'].fill = fill_kpi
    ws['B5'].border = thin_border
    ws['B5'].alignment = Alignment(horizontal='center')
    
    # Write Concentration Section Table (Row 8 Onwards)
    ws['A7'] = 'SYSTEMIC CONCENTRATION ANALYSIS (BY TICKER)'
    ws['A7'].font = font_section
    
    headers = ['Ticker Symbol', 'Total Outstanding Incidents', 'Concentrated Risk Value Exposure']
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=8, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    current_row = 9
    for _, row in df_concentration.iterrows():
        ws.cell(row=current_row, column=1, value=row['Ticker']).font = font_bold
        ws.cell(row=current_row, column=2, value=int(row['Total_Incidents'])).number_format = '#,##0'
        ws.cell(row=current_row, column=3, value=float(row['Concentrated_Risk'])).number_format = '0,##0.00'
        
        # Format alignment and layout styles
        for c_idx in range(1, 4):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.font = font_data if c_idx != 1 else font_bold
            cell.border = thin_border
            cell.fill = fill_row_break
            if c_idx in (2, 3):
                cell.alignment = Alignment(horizontal='right')
        current_row += 1
        
    # Set explicit custom column boundaries
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 32
    
    wb.save(output_path)
    print(f'[SUCCESS] SQL Management Summary Report cleanly saved to: {output_path}')

if __name__ == "__main__":
    generate_sql_management_report()
